from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Dict, Any, Optional
import uuid
from datetime import datetime
import re

from groq import Groq
# NLP and ML imports
import spacy
import nltk
from nltk.corpus import wordnet
from sentence_transformers import SentenceTransformer
from textblob import TextBlob
import numpy as np
from sklearn.metrics.pairwise import cosine_similarity
import networkx as nx

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")


GROQ_API_KEY = os.getenv("GROQ_API_KEY")
# Initialize NLP models
try:
    nlp = spacy.load("en_core_web_sm")
except OSError:
    print("SpaCy model not found. Install with: python -m spacy download en_core_web_sm")
    nlp = None

# Initialize sentence transformer
try:
    sentence_model = SentenceTransformer('all-MiniLM-L6-v2')
except:
    print("Sentence transformer not available")
    sentence_model = None

# Download required NLTK data
try:
    nltk.download('wordnet', quiet=True)
    nltk.download('punkt', quiet=True)
    nltk.download('averaged_perceptron_tagger', quiet=True)
    nltk.download('stopwords', quiet=True)
except:
    pass

# Define Models
class Requirement(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    text: str
    timestamp: datetime = Field(default_factory=datetime.utcnow)

class Proposition(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    text: str
    requirement_id: str
    label: str  # P1, P2, etc.
    type: str  # software behavior, condition, state change
    negations: List[str] = []
    best_negation: str = ""
    can_coexist_with_negation: bool = True
    quality_score: Optional[float] = 0
    quality_scores: Dict[str, Any] = {}
    detected_issues: List[str] = []

class LogicalRelation(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    prop1_id: str
    prop2_id: str
    relation_type: str  # dependent, biconditional, disjoint, contradictory
    confidence: float
    explanation: str

class AnalysisResult(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    requirement_id: str
    propositions: List[Proposition]
    relations: List[LogicalRelation]
    overall_stats: Dict[str, Any]
    timestamp: datetime = Field(default_factory=datetime.utcnow)

class RequirementInput(BaseModel):
    text: str

# Vague terms dictionary
VAGUE_TERMS = {
    'fast', 'slow', 'quick', 'quickly', 'soon', 'later', 'better', 'worse',
    'good', 'bad', 'high', 'low', 'many', 'few', 'some', 'several',
    'appropriate', 'suitable', 'adequate', 'reasonable', 'efficient', 'effective',
    'comfortable', 'convenient', 'user-friendly', 'easy', 'difficult',
    'large', 'small', 'big', 'little', 'significant', 'considerable'
}

class RequirementProcessor:
    def __init__(self):
        self.nlp = nlp
        self.sentence_model = sentence_model
        
    def extract_propositions(self, requirement_text: str) -> List[Dict[str, Any]]:
        """Extract minimal proposition units from requirement text"""
        if not self.nlp:
            # Fallback simple extraction
            sentences = re.split(r'[.!?]+', requirement_text)
            propositions = []
            for i, sent in enumerate(sentences):
                if sent.strip():
                    subprops = re.split(r'\b(?:and|or|but|nor|yet|so|when)\b', sent.strip(), flags=re.IGNORECASE)
                    for clause in subprops:
                        if clause.strip():
                            propositions.append({
                                'text': clause.strip(),
                                'label': f'P{len(propositions)+1}',
                                'type': self._classify_proposition_type(clause.strip())
                            })
            return propositions

        doc = self.nlp(requirement_text)
        propositions = []
        prop_counter = 1

        # Split using spaCy sentences, then refine with regex
        for sent in doc.sents:
            parts = re.split(r'\b(?:and|or|but|nor|yet|so)\b', sent.text, flags=re.IGNORECASE)
            for clause in parts:
                clean = clause.strip()
                if clean and len(clean) > 5:
                    propositions.append({
                        'text': clean,
                        'label': f'P{prop_counter}',
                        'type': self._classify_proposition_type(clean)
                    })
                    prop_counter += 1

        return propositions

    
    def _split_on_conjunctions(self, text: str) -> List[str]:
        """Split text on coordinating conjunctions"""
        # Split on 'and', 'or', 'but', 'if', 'when', 'while', 'once'
        conjunctions = r'\b(?:and|or|but|if|when|while|once)\b'
        parts = re.split(conjunctions, text, flags=re.IGNORECASE)
        return [part.strip() for part in parts if part.strip()]
    
    def _classify_proposition_type(self, text: str) -> str:
        text_lower = text.lower()

        condition_keywords = {'if', 'when', 'while', 'unless', 'provided that'}
        state_change_keywords = {'turn on', 'turn off', 'start', 'stop', 'change', 'enable', 'disable'}
        behavior_keywords = {'shall', 'must', 'should', 'will'}

        # Match condition phrases first
        if any(kw in text_lower for kw in condition_keywords):
            return 'condition'

        # Then look for state-related actions
        if any(kw in text_lower for kw in state_change_keywords):
            return 'state change'

        # Finally, general software behavior
        if any(kw in text_lower for kw in behavior_keywords):
            return 'software behavior'

        # Fallback
        return 'unspecified'

    
    def generate_negations(self, text: str) -> List[str]:
        groq_client = Groq(api_key=GROQ_API_KEY)
        if not text.strip():
            return []

        prompt = f"""
    You are an AI trained to help requirement engineers negate statements.

    Given the following requirement, generate two logically correct negated versions. Focus on preserving grammar, clarity, and flipping the logic meaningfully.
    Return only the negated statements and nothing else with it, no labelling.
    Requirement: "{text}"
    Negations:
    """

        try:
            chat_completion = groq_client.chat.completions.create(
                model="llama3-70b-8192",
                messages=[{"role": "user", "content": prompt}],
                temperature=0.3
            )
            response = chat_completion.choices[0].message.content.strip()

            # Split negations by lines
            negs = [line.strip("- ").strip() for line in response.split('\n') if line.strip()]
            return negs[:3]

        except Exception as e:
            print("Groq negation error:", e)
            return []


    
    def _get_antonyms(self, word: str) -> List[str]:
        """Get antonyms using WordNet"""
        antonyms = []
        try:
            for syn in wordnet.synsets(word):
                for lemma in syn.lemmas():
                    if lemma.antonyms():
                        antonyms.extend([ant.name() for ant in lemma.antonyms()])
        except:
            pass
        return list(set(antonyms))
    
    def select_best_negation(self, original: str, negations: List[str]) -> str:
        """Select the most semantically relevant negation"""
        if not negations or not self.sentence_model:
            return negations[0] if negations else f"NOT({original})"
        
        try:
            original_embedding = self.sentence_model.encode([original])
            negation_embeddings = self.sentence_model.encode(negations)
            
            # Choose negation with highest semantic similarity (but still opposite meaning)
            similarities = cosine_similarity(original_embedding, negation_embeddings)[0]
            best_idx = np.argmax(similarities)
            return negations[best_idx]
        except:
            return negations[0] if negations else f"NOT({original})"
    
    def check_coexistence(self, original: str, negation: str) -> bool:
        """Check if original and negated propositions can coexist logically"""
        # Simple logical inference - if they're direct opposites, they can't coexist
        if 'not' in negation.lower() and self._extract_core_action(original) in negation:
            return False
        
        # Check for contradictory actions
        contradictory_pairs = [
            ('turn on', 'turn off'), ('start', 'stop'), ('enable', 'disable'),
            ('allow', 'deny'), ('grant', 'revoke'), ('connect', 'disconnect')
        ]
        
        orig_lower = original.lower()
        neg_lower = negation.lower()
        
        for pair in contradictory_pairs:
            if pair[0] in orig_lower and pair[1] in neg_lower:
                return False
            if pair[1] in orig_lower and pair[0] in neg_lower:
                return False
        
        return True
    
    def _extract_core_action(self, text: str) -> str:
        """Extract the core action from a proposition"""
        if not self.nlp:
            return text
        
        doc = self.nlp(text)
        for token in doc:
            if token.pos_ == 'VERB' and token.dep_ in ['ROOT', 'ccomp']:
                return token.lemma_
        return text
    
    def detect_logical_relations(self, propositions: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Detect logical relationships between propositions"""
        relations = []
        
        for i, prop1 in enumerate(propositions):
            for j, prop2 in enumerate(propositions):
                if i >= j:  # Avoid duplicates and self-comparison
                    continue
                
                relation = self._infer_relation(prop1, prop2)
                if relation:
                    relations.append(relation)
        
        return relations
    
    def _infer_relation(self, prop1: Dict[str, Any], prop2: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """Infer logical relation between two propositions"""
        text1 = prop1['text'].lower()
        text2 = prop2['text'].lower()
        
        # Check for contradictory relations
        if self._are_contradictory(text1, text2):
            return {
                'prop1_id': prop1.get('id', ''),
                'prop2_id': prop2.get('id', ''),
                'relation_type': 'contradictory',
                'confidence': 0.8,
                'explanation': f"{prop1['label']} contradicts {prop2['label']} because they propose opposite outcomes for the same condition."
            }
        
        # Check for dependent relations (conditional)
        if self._are_dependent(text1, text2):
            return {
                'prop1_id': prop1.get('id', ''),
                'prop2_id': prop2.get('id', ''),
                'relation_type': 'dependent',
                'confidence': 0.7,
                'explanation': f"{prop1['label']} depends on {prop2['label']} as it contains conditional logic."
            }
        
        # Check for biconditional relations
        if self._are_biconditional(text1, text2):
            return {
                'prop1_id': prop1.get('id', ''),
                'prop2_id': prop2.get('id', ''),
                'relation_type': 'biconditional',
                'confidence': 0.6,
                'explanation': f"{prop1['label']} and {prop2['label']} have a biconditional relationship (if and only if)."
            }
        
        # Check semantic similarity for relatedness
        if self.sentence_model:
            try:
                embeddings = self.sentence_model.encode([text1, text2])
                similarity = cosine_similarity([embeddings[0]], [embeddings[1]])[0][0]
                
                if similarity > 0.5:
                    return {
                        'prop1_id': prop1.get('id', ''),
                        'prop2_id': prop2.get('id', ''),
                        'relation_type': 'dependent',
                        'confidence': similarity,
                        'explanation': f"{prop1['label']} and {prop2['label']} are semantically related."
                    }
                else:
                    return {
                        'prop1_id': prop1.get('id', ''),
                        'prop2_id': prop2.get('id', ''),
                        'relation_type': 'disjoint',
                        'confidence': 1 - similarity,
                        'explanation': f"{prop1['label']} and {prop2['label']} are logically unrelated."
                    }
            except:
                pass
        
        return None
    
    def _are_contradictory(self, text1: str, text2: str) -> bool:
        """Check if two propositions are contradictory"""
        # Check for explicit negations
        if ('not' in text1 and 'not' not in text2) or ('not' in text2 and 'not' not in text1):
            # Extract core content without negation
            core1 = text1.replace('not', '').strip()
            core2 = text2.replace('not', '').strip()
            if self._texts_similar(core1, core2):
                return True
        
        # Check for contradictory action pairs
        contradictory_pairs = [
            ('turn on', 'turn off'), ('start', 'stop'), ('enable', 'disable'),
            ('allow', 'deny'), ('grant', 'revoke'), ('connect', 'disconnect')
        ]
        
        for pair in contradictory_pairs:
            if (pair[0] in text1 and pair[1] in text2) or (pair[1] in text1 and pair[0] in text2):
                return True
        
        return False
    
    def _are_dependent(self, text1: str, text2: str) -> bool:
        """Check if propositions have dependent relationship"""
        conditional_words = ['if', 'when', 'while', 'once', 'after', 'before', 'unless']
        return any(word in text1 or word in text2 for word in conditional_words)
    
    def _are_biconditional(self, text1: str, text2: str) -> bool:
        """Check if propositions have biconditional relationship"""
        # Look for paired actions like on/off, start/stop
        biconditional_indicators = [
            ('turn on', 'turn off'), ('start', 'stop'), ('open', 'close'),
            ('connect', 'disconnect'), ('enable', 'disable')
        ]
        
        for pair in biconditional_indicators:
            if (pair[0] in text1 and pair[1] in text2) or (pair[1] in text1 and pair[0] in text2):
                return True
        
        return False
    
    def _texts_similar(self, text1: str, text2: str) -> bool:
        """Check if two texts are similar in meaning"""
        if not self.sentence_model:
            return text1.strip() == text2.strip()
        
        try:
            embeddings = self.sentence_model.encode([text1, text2])
            similarity = cosine_similarity([embeddings[0]], [embeddings[1]])[0][0]
            return similarity > 0.7
        except:
            return False
    
    def analyze_quality(self, proposition_text: str) -> Dict[str, Any]:
        """Analyze quality dimensions of a proposition"""
        quality_scores = {
            'completeness': self._check_completeness(proposition_text),
            'correctness': self._check_correctness(proposition_text),
            'clarity': self._check_clarity(proposition_text),
            'explicitness': self._check_explicitness(proposition_text)
        }
        
        detected_issues = []
        
        # Check for ambiguity
        if self._detect_ambiguity(proposition_text):
            detected_issues.append('ambiguity')
        
        # Check for vagueness
        if self._detect_vagueness(proposition_text):
            detected_issues.append('vagueness')
        
        # Check for inconsistency (placeholder for now)
        if self._detect_inconsistency(proposition_text):
            detected_issues.append('inconsistency')
        
        return {
            'quality_scores': quality_scores,
            'detected_issues': detected_issues,
            'overall_score': sum(quality_scores.values()) / len(quality_scores)
        }
    
    def _check_completeness(self, text: str) -> float:
        """Check if proposition is complete"""
        # Simple heuristic: longer, more detailed propositions are more complete
        word_count = len(text.split())
        has_subject = bool(re.search(r'\b(system|user|application|heater|robot)\b', text.lower()))
        has_action = bool(re.search(r'\b(shall|will|must|turn|move|notify)\b', text.lower()))
        has_condition = bool(re.search(r'\b(if|when|while|once)\b', text.lower()))
        
        score = 0.0
        if word_count >= 8: score += 0.3
        if has_subject: score += 0.3
        if has_action: score += 0.3
        if has_condition: score += 0.1
        
        return min(score, 1.0)
    
    def _check_correctness(self, text: str) -> float:
        """Check logical and factual validity"""
        # Basic checks for logical consistency
        if 'and' in text and 'or' in text:
            return 0.7  # Mixed logical operators might be confusing
        return 0.9  # Assume mostly correct for now
    
    def _check_clarity(self, text: str) -> float:
        """Check if proposition is clearly worded"""
        unclear_terms = ['thing', 'stuff', 'something', 'it', 'this', 'that']
        penalty = sum(1 for term in unclear_terms if term in text.lower()) * 0.1
        return max(0.9 - penalty, 0.0)
    
    def _check_explicitness(self, text: str) -> float:
        """Check if all actions, conditions, and actors are clearly mentioned"""
        has_explicit_actor = bool(re.search(r'\b(system|user|application|heater|robot|the \w+)\b', text.lower()))
        has_explicit_action = bool(re.search(r'\b(notify|turn on|turn off|move|avoid|reach)\b', text.lower()))
        
        score = 0.5
        if has_explicit_actor: score += 0.25
        if has_explicit_action: score += 0.25
        
        return score
    
    def _detect_ambiguity(self, text: str) -> bool:
        """Detect ambiguous terms or references"""
        ambiguous_terms = ['it', 'this', 'that', 'they', 'them', 'something', 'anything']
        return any(term in text.lower().split() for term in ambiguous_terms)
    
    def _detect_vagueness(self, text: str) -> bool:
        """Detect vague terms"""
        return any(term in text.lower() for term in VAGUE_TERMS)
    
    def _detect_inconsistency(self, text: str) -> bool:
        """Detect internal inconsistencies"""
        # Simple check for contradictory statements within same proposition
        if 'shall' in text and 'shall not' in text:
            return True
        return False
    def propagate_subject_to_clauses(self, clauses: List[Dict[str, str]]) -> List[Dict[str, str]]:
        """Ensure that each clause includes the subject if missing."""
        if not clauses:
            return []

        # Naively extract subject from the first clause (could enhance with NLP)
        first_clause = clauses[0]['text']
        match = re.match(r'^([A-Z][a-z0-9\s]+?)\s+(shall|will|must|can|should)\b', first_clause, re.IGNORECASE)
        subject = match.group(1) if match else None

        if not subject:
            return clauses  # Cannot find a subject, return as-is

        updated = []
        for clause in clauses:
            text = clause['text'].strip()
            if re.match(r'^(shall|will|must|can|should)\b', text, re.IGNORECASE):
                clause['text'] = f"{subject} {text}"
            updated.append(clause)

        return updated


# Initialize processor
processor = RequirementProcessor()

# API Routes
@api_router.post("/analyze-requirement", response_model=AnalysisResult)
async def analyze_requirement(req_input: RequirementInput):
    """Analyze a single requirement and extract propositions with relationships"""
    # print(RequirementInput.text)
    try:
        # Create requirement record
        requirement = Requirement(text=req_input.text)
        print(requirement)
        await db.requirements.insert_one(requirement.dict())

        # Extract propositions
        raw_propositions = processor.extract_propositions(req_input.text)
        raw_propositions =processor.propagate_subject_to_clauses(raw_propositions)
        print(raw_propositions)
        propositions = []
        
        for raw_prop in raw_propositions:
            # Generate negations
            negations = processor.generate_negations(raw_prop['text'])
            best_negation = processor.select_best_negation(raw_prop['text'], negations)
            can_coexist = processor.check_coexistence(raw_prop['text'], best_negation)

            # Analyze quality
            quality_analysis = processor.analyze_quality(raw_prop['text'])
            quality_scores = quality_analysis['quality_scores']
            detected_issues = quality_analysis['detected_issues']

            prop = Proposition(
                text=raw_prop['text'],
                requirement_id=requirement.id,
                label=raw_prop['label'],
                type=raw_prop['type'],
                negations=negations,
                best_negation=best_negation,
                can_coexist_with_negation=can_coexist,
                quality_scores=quality_scores,
                detected_issues=detected_issues
            )

            # ✅ Inject flat average score
            prop.quality_score = (
                sum(quality_scores.values()) / len(quality_scores)
                if quality_scores else 0
            )

            propositions.append(prop)

        # Detect logical relations
        raw_relations = processor.detect_logical_relations([p.dict() for p in propositions])
        relations = [LogicalRelation(**rel) for rel in raw_relations]

        # Calculate overall statistics
        total_props = len(propositions)
        ambiguous_count = sum(1 for p in propositions if 'ambiguity' in p.detected_issues)
        vague_count = sum(1 for p in propositions if 'vagueness' in p.detected_issues)
        inconsistent_count = sum(1 for p in propositions if 'inconsistency' in p.detected_issues)

        overall_stats = {
            'total_propositions': total_props,
            'ambiguous_propositions': ambiguous_count,
            'vague_propositions': vague_count,
            'inconsistent_propositions': inconsistent_count,
            'total_relations': len(relations),
            'ambiguity_percentage': (ambiguous_count / total_props * 100) if total_props > 0 else 0,
            'average_quality_score': (
                sum(p.quality_score for p in propositions) / total_props
                if total_props > 0 else 0
            )
        }

        result = AnalysisResult(
            requirement_id=requirement.id,
            propositions=propositions,
            relations=relations,
            overall_stats=overall_stats
        )

        # Save analysis result
        await db.analysis_results.insert_one(result.dict())

        return result

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Analysis failed: {str(e)}")


@api_router.post("/analyze-bulk")
async def analyze_bulk_requirements(file: UploadFile = File(...)):
    """Analyze multiple requirements from uploaded text file"""
    try:
        content = await file.read()
        text = content.decode('utf-8')
        
        # Split into individual requirements (assuming one per line or paragraph)
        requirements = [req.strip() for req in text.split('\n') if req.strip()]
        
        results = []
        for req_text in requirements:
            if req_text:
                req_input = RequirementInput(text=req_text)
                result = await analyze_requirement(req_input)
                results.append(result)
        
        return {"results": results, "count": len(results)}
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Bulk analysis failed: {str(e)}")

@api_router.get("/analysis-results", response_model=List[AnalysisResult])
async def get_analysis_results():
    """Get all analysis results"""
    try:
        results = await db.analysis_results.find().to_list(1000)
        return [AnalysisResult(**result) for result in results]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to fetch results: {str(e)}")

@api_router.get("/analysis-results/{result_id}", response_model=AnalysisResult)
async def get_analysis_result(result_id: str):
    """Get specific analysis result"""
    try:
        result = await db.analysis_results.find_one({"id": result_id})
        if not result:
            raise HTTPException(status_code=404, detail="Analysis result not found")
        return AnalysisResult(**result)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to fetch result: {str(e)}")

@api_router.delete("/analysis-results/{result_id}")
async def delete_analysis_result(result_id: str):
    """Delete analysis result"""
    try:
        result = await db.analysis_results.delete_one({"id": result_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Analysis result not found")
        return {"message": "Analysis result deleted successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to delete result: {str(e)}")
    


@api_router.post("/chat")
async def chat_with_groq(payload: dict):
    gclient = Groq(api_key=GROQ_API_KEY)  
    text = payload.get("text", "").strip()
    query = payload.get("query", "").strip()

    if not text or not query:
        raise HTTPException(status_code=400, detail="Both text and query are required.")

    prompt = f"""
You are a helpful assistant for requirement engineers. Analyze the following requirement text:

{text}

User's question: {query}
Be clear, concise, and precise.
"""

    try:
        chat_completion = gclient.chat.completions.create(
            model="llama3-70b-8192",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.3
        )

        reply = chat_completion.choices[0].message.content
        return {"response": reply}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Groq API failed: {str(e)}")


from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint
from io import BytesIO

@api_router.post("/report")
async def generate_report(payload: dict):
    try:
        buffer = BytesIO()
        wb = Workbook()

        # === Sheet: Dashboard ===
        stats = payload.get("stats", {})
        props = payload.get("propositions", [])
        issue_count: dict = {}

        for p in props:
            for issue in p.get("detected_issues", []):
                issue_count[issue] = issue_count.get(issue, 0) + 1

        dashboard_ws = wb.active
        dashboard_ws.title = "Dashboard"
        dashboard_ws.append(["Metric", "Value"])
        for k, v in stats.items():
            dashboard_ws.append([k.replace('_', ' ').capitalize(), round(v, 2) if isinstance(v, (int, float)) else v])
        for cell in dashboard_ws["1:1"]:
            cell.font = Font(bold=True)


        # === Sheet: Propositions ===
        prop_ws = wb.create_sheet("Propositions")
        prop_ws.append(["Label", "Text", "Quality Score", "Issues", "Best Negation", "Can Coexist"])
        for p in props:
            prop_ws.append([
                p.get("label", ""),
                p.get("text", ""),
                round(p.get("quality_score", 0) * 100),
                ", ".join(p.get("detected_issues", [])),
                p.get("best_negation", ""),
                "Yes" if p.get("can_coexist_with_negation") else "No"
            ])
        for cell in prop_ws["1:1"]:
            cell.font = Font(bold=True)

        # === Sheet: Relations ===
        rels = payload.get("relations", [])
        rel_ws = wb.create_sheet("Relations")
        rel_ws.append([ "Relation Type", "Confidence", "Explanation"])

        for r in rels:
            rel_ws.append([
                r.get("relation_type", ""),
                round(r.get("confidence", 0) * 100),
                r.get("explanation", "")
            ])
        for cell in rel_ws["1:1"]:
            cell.font = Font(bold=True)

        # Save to buffer and return
        wb.save(buffer)
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=requirement_report.xlsx"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate Excel report: {str(e)}")



@api_router.get("/")
async def root():
    return {"message": "Requirement Elicitation Tool API"}

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()